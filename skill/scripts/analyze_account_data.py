#!/usr/bin/env python
"""Summarize Xiaohongshu/Douyin CSV or XLSX exports into JSON, Markdown, and HTML."""

from __future__ import annotations

import argparse
import csv
import html
import json
import math
import re
from collections import defaultdict
from datetime import date
from pathlib import Path
from statistics import median
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


FIELD_ALIASES = {
    "title": ["笔记标题", "作品标题", "标题", "内容标题", "视频标题", "名称"],
    "published_at": ["发布时间", "首次发布时间", "发布日期", "创建时间", "时间", "发布于"],
    "content_type": ["体裁", "内容类型", "作品类型", "形式", "类型"],
    "impressions": ["曝光", "曝光量", "展现", "展现量", "推荐曝光", "展示量"],
    "views": ["阅读", "阅读量", "观看", "观看量", "播放", "播放量", "浏览", "浏览量", "视频播放量"],
    "likes": ["点赞", "点赞数", "赞"],
    "favorites": ["收藏", "收藏数", "保存", "保存数"],
    "comments": ["评论", "评论数"],
    "shares": ["分享", "分享数", "转发", "转发数"],
    "follows": ["涨粉", "新增粉丝", "粉丝增长", "关注", "新增关注"],
    "ctr": ["点击率", "阅读率", "封面点击率", "ctr", "CTR"],
    "completion_rate": ["完播率", "播放完成率", "completion"],
    "avg_watch_seconds": ["人均观看时长", "平均观看时长", "平均播放时长", "avg watch"],
}

DEFAULT_PILLARS = {
    "AI工具效率": ["ai", "chatgpt", "claude", "gemini", "vibe", "agent", "智能体", "工具", "效率", "自动化", "提效"],
    "案例作品视觉": ["案例", "作品", "视觉", "设计", "海报", "短片", "视频", "品牌", "logo", "包装"],
    "教程方法论": ["教程", "怎么", "如何", "方法", "步骤", "入门", "攻略", "拆解", "复盘"],
    "行业趋势观点": ["趋势", "行业", "未来", "时代", "洞察", "观点", "认知", "机会"],
    "工作室幕后成长": ["工作室", "创业", "团队", "接单", "客户", "报价", "成长", "幕后"],
    "商业转化服务": ["服务", "报价", "咨询", "方案", "客户", "项目", "交付", "品牌"],
}


def norm(text: Any) -> str:
    return re.sub(r"\s+", "", str(text or "")).lower()


def parse_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    multiplier = 1.0
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError:
            return 0.0
    if text.endswith("万"):
        multiplier = 10000.0
        text = text[:-1]
    elif text.endswith("k") or text.endswith("K"):
        multiplier = 1000.0
        text = text[:-1]
    try:
        return float(text) * multiplier
    except ValueError:
        return 0.0


def safe_div(a: float, b: float) -> float:
    return a / b if b else 0.0


def read_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() in {".csv", ".txt"}:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise RuntimeError("openpyxl is required for XLSX files.")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header_index = detect_header_row(rows)
        headers = [str(h or "").strip() for h in rows[header_index]]
        return [
            {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            for row in rows[header_index + 1 :]
            if any(cell not in (None, "") for cell in row)
        ]
    raise ValueError(f"Unsupported file type: {path.suffix}")


def detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    expected = {norm(alias) for aliases in FIELD_ALIASES.values() for alias in aliases}
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:10]):
        cells = [norm(cell) for cell in row]
        score = 0
        for cell in cells:
            if cell in expected:
                score += 2
            elif any(alias in cell or cell in alias for alias in expected):
                score += 1
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def map_fields(headers: list[str]) -> dict[str, str]:
    normalized = {norm(h): h for h in headers}
    result: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            key = norm(alias)
            if key in normalized:
                result[field] = normalized[key]
                break
        if field not in result:
            for key, original in normalized.items():
                if any(norm(alias) in key for alias in aliases):
                    result[field] = original
                    break
    return result


def load_pillars(path: Path | None) -> dict[str, list[str]]:
    if not path:
        return DEFAULT_PILLARS
    data = json.loads(path.read_text(encoding="utf-8"))
    return {str(k): [str(x).lower() for x in v] for k, v in data.items()}


def load_trends(path: Path | None) -> list[str]:
    if not path:
        return []
    if path.suffix.lower() == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data = data.get("trends", [])
        return [str(x).strip() for x in data if str(x).strip()]
    return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def classify_pillar(title: str, pillars: dict[str, list[str]]) -> str:
    lowered = title.lower()
    scores = {}
    for pillar, keywords in pillars.items():
        scores[pillar] = sum(1 for keyword in keywords if keyword and keyword in lowered)
    best = max(scores, key=scores.get) if scores else "其他"
    return best if scores.get(best, 0) > 0 else "其他"


def enrich_rows(rows: list[dict[str, Any]], field_map: dict[str, str], pillars: dict[str, list[str]], trends: list[str]) -> list[dict[str, Any]]:
    enriched = []
    for row in rows:
        title = str(row.get(field_map.get("title", ""), "") or "").strip()
        content_type = str(row.get(field_map.get("content_type", ""), "") or "").strip()
        impressions = parse_number(row.get(field_map.get("impressions", "")))
        views = parse_number(row.get(field_map.get("views", "")))
        likes = parse_number(row.get(field_map.get("likes", "")))
        favorites = parse_number(row.get(field_map.get("favorites", "")))
        comments = parse_number(row.get(field_map.get("comments", "")))
        shares = parse_number(row.get(field_map.get("shares", "")))
        follows = parse_number(row.get(field_map.get("follows", "")))
        ctr_raw = parse_number(row.get(field_map.get("ctr", "")))
        completion_rate = parse_number(row.get(field_map.get("completion_rate", "")))
        avg_watch_seconds = parse_number(row.get(field_map.get("avg_watch_seconds", "")))
        engagement = likes + favorites + comments + shares
        base_views = views or impressions
        trend_hits = [t for t in trends if t.lower() in title.lower()]
        enriched.append(
            {
                "title": title or "未命名内容",
                "published_at": str(row.get(field_map.get("published_at", ""), "") or ""),
                "content_type": content_type,
                "pillar": classify_pillar(title, pillars),
                "impressions": impressions,
                "views": views,
                "likes": likes,
                "favorites": favorites,
                "comments": comments,
                "shares": shares,
                "follows": follows,
                "ctr": ctr_raw or safe_div(views, impressions),
                "completion_rate": completion_rate,
                "avg_watch_seconds": avg_watch_seconds,
                "engagement": engagement,
                "interaction_rate": safe_div(engagement, base_views),
                "utility_ratio": safe_div(favorites, likes),
                "follows_per_1000": safe_div(follows, base_views) * 1000,
                "trend_hits": trend_hits,
            }
        )
    return enriched


def summarize(enriched: list[dict[str, Any]], platform: str, account: str, source: Path, field_map: dict[str, str]) -> dict[str, Any]:
    totals = defaultdict(float)
    for row in enriched:
        for key in ["impressions", "views", "likes", "favorites", "comments", "shares", "follows", "engagement"]:
            totals[key] += row[key]
    pillar_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in enriched:
        pillar_rows[row["pillar"]].append(row)
    pillar_summary = []
    total_reach = sum(r["views"] or r["impressions"] for r in enriched)
    total_follows = sum(r["follows"] for r in enriched)
    overall_follows_per_1000 = safe_div(total_follows, total_reach) * 1000
    overall_interaction = safe_div(sum(r["engagement"] for r in enriched), total_reach)
    for pillar, rows in sorted(pillar_rows.items(), key=lambda kv: sum(r["views"] or r["impressions"] for r in kv[1]), reverse=True):
        reach = sum(r["views"] or r["impressions"] for r in rows)
        engagement = sum(r["engagement"] for r in rows)
        follows = sum(r["follows"] for r in rows)
        saves = sum(r["favorites"] for r in rows)
        likes = sum(r["likes"] for r in rows)
        pillar_summary.append(
            {
                "pillar": pillar,
                "posts": len(rows),
                "reach": reach,
                "engagement": engagement,
                "interaction_rate": safe_div(engagement, reach),
                "utility_ratio": safe_div(saves, likes),
                "follows_per_1000": safe_div(follows, reach) * 1000,
                "decision": decide_pillar(
                    reach,
                    engagement,
                    follows,
                    saves,
                    likes,
                    total_reach,
                    overall_follows_per_1000,
                    overall_interaction,
                ),
            }
        )
    reach_values = [r["views"] or r["impressions"] for r in enriched]
    median_reach = median(reach_values) if reach_values else 0
    top_posts = sorted(enriched, key=lambda r: (r["views"] or r["impressions"], r["engagement"]), reverse=True)[:10]
    hidden_gems = sorted(
        [r for r in enriched if (r["views"] or r["impressions"]) <= max(median_reach, 1)],
        key=lambda r: (r["follows_per_1000"], r["utility_ratio"], r["interaction_rate"]),
        reverse=True,
    )[:8]
    return {
        "account": account,
        "platform": platform,
        "source": str(source),
        "generated_at": date.today().isoformat(),
        "post_count": len(enriched),
        "totals": dict(totals),
        "averages": {
            "ctr": mean([r["ctr"] for r in enriched]),
            "interaction_rate": mean([r["interaction_rate"] for r in enriched]),
            "utility_ratio": safe_div(totals["favorites"], totals["likes"]),
            "follows_per_1000": safe_div(totals["follows"], totals["views"] or totals["impressions"]) * 1000,
        },
        "pillars": pillar_summary,
        "top_posts": top_posts,
        "hidden_gems": hidden_gems,
        "data_gaps": find_data_gaps(enriched),
        "field_map": field_map,
    }


def mean(values: list[float]) -> float:
    values = [v for v in values if not math.isnan(v)]
    return sum(values) / len(values) if values else 0.0


def decide_pillar(
    reach: float,
    engagement: float,
    follows: float,
    saves: float,
    likes: float,
    total_reach: float,
    overall_follows_per_1000: float,
    overall_interaction: float,
) -> str:
    utility = safe_div(saves, likes)
    follow_rate = safe_div(follows, reach) * 1000
    interaction = safe_div(engagement, reach)
    reach_share = safe_div(reach, total_reach)
    if reach_share >= 0.45 and follow_rate < overall_follows_per_1000 * 0.6:
        return "保留流量，补强转粉承接"
    if follow_rate >= max(overall_follows_per_1000 * 1.8, 8) and reach_share < 0.2:
        return "加大包装测试"
    if interaction >= overall_interaction * 1.2 and utility >= 0.5:
        return "系列化放大"
    if interaction < max(overall_interaction * 0.55, 0.03) and follow_rate < overall_follows_per_1000:
        return "收窄或暂停"
    if utility >= 0.8:
        return "做收藏型栏目"
    return "稳态小步测试"


def find_data_gaps(enriched: list[dict[str, Any]]) -> list[str]:
    gaps = []
    for field, label in [
        ("impressions", "曝光/展现"),
        ("views", "阅读/播放"),
        ("follows", "涨粉"),
        ("avg_watch_seconds", "人均观看时长"),
    ]:
        if not any(r.get(field) for r in enriched):
            gaps.append(label)
    return gaps


def fmt_num(value: float) -> str:
    if abs(value) >= 10000:
        return f"{value / 10000:.1f}w"
    return f"{value:.0f}"


def fmt_pct(value: float) -> str:
    return f"{value * 100:.2f}%"


def write_json(summary: dict[str, Any], path: Path) -> None:
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def write_markdown(summary: dict[str, Any], path: Path) -> None:
    lines = [
        f"# {summary['account']} {summary['platform']}账号分析草稿",
        "",
        f"- 数据源: `{summary['source']}`",
        f"- 内容数: {summary['post_count']}",
        f"- 总互动: {fmt_num(summary['totals'].get('engagement', 0))}",
        f"- 收藏/点赞比: {summary['averages']['utility_ratio']:.2f}",
        f"- 千次阅读/播放涨粉: {summary['averages']['follows_per_1000']:.2f}",
        "",
        "## 板块判断",
        "",
    ]
    for p in summary["pillars"]:
        lines.append(
            f"- **{p['pillar']}**: {p['posts']}篇, 触达{fmt_num(p['reach'])}, "
            f"互动率{fmt_pct(p['interaction_rate'])}, 收藏/点赞{p['utility_ratio']:.2f}, "
            f"千次涨粉{p['follows_per_1000']:.2f}。建议: {p['decision']}。"
        )
    lines.extend(["", "## Top内容", ""])
    for post in summary["top_posts"][:5]:
        lines.append(f"- {post['title']} | {post['pillar']} | 触达{fmt_num(post['views'] or post['impressions'])} | 互动{fmt_num(post['engagement'])}")
    if summary["data_gaps"]:
        lines.extend(["", "## 数据缺口", "", "- " + "、".join(summary["data_gaps"])])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_html(summary: dict[str, Any], path: Path) -> None:
    pillars = summary["pillars"]
    max_reach = max([p["reach"] for p in pillars] or [1])
    pillar_rows = "\n".join(
        f"""
        <div class="bar-row">
          <div>{html.escape(p['pillar'])}</div>
          <div class="bar"><span style="width:{max(4, p['reach'] / max_reach * 100):.1f}%"></span></div>
          <div>{fmt_num(p['reach'])}</div>
        </div>
        """
        for p in pillars
    )
    decisions = "\n".join(
        f"""
        <div class="pill">
          <span class="tag">{html.escape(p['decision'])}</span>
          <strong>{html.escape(p['pillar'])}</strong>
          <div>互动率 {fmt_pct(p['interaction_rate'])} · 收藏/点赞 {p['utility_ratio']:.2f} · 千次涨粉 {p['follows_per_1000']:.2f}</div>
        </div>
        """
        for p in pillars[:6]
    )
    next_posts = "\n".join(
        f"""
        <div class="next-item">
          <strong>{html.escape(post['title'])}</strong>
          <div>{html.escape(post['pillar'])} · 触达 {fmt_num(post['views'] or post['impressions'])} · 互动 {fmt_num(post['engagement'])}</div>
        </div>
        """
        for post in summary["hidden_gems"][:5]
    )
    html_doc = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(summary['account'])}账号分析</title>
  <link rel="stylesheet" href="dashboard-theme.css">
</head>
<body>
  <main class="dashboard">
    <section class="topbar">
      <div>
        <h1 class="title">{html.escape(summary['account'])}账号分析</h1>
        <p class="subtitle">{html.escape(summary['platform'])} · {summary['post_count']}条内容 · {summary['generated_at']}</p>
      </div>
      <span class="tag">Dopamine Flat Dashboard</span>
    </section>
    <section class="metric-grid">
      <div class="metric"><div class="metric-label">TOTAL REACH</div><div class="metric-value">{fmt_num(summary['totals'].get('views') or summary['totals'].get('impressions', 0))}</div><div class="metric-note">阅读/播放优先</div></div>
      <div class="metric"><div class="metric-label">AVG CTR</div><div class="metric-value">{fmt_pct(summary['averages']['ctr'])}</div><div class="metric-note">封面与标题效率</div></div>
      <div class="metric"><div class="metric-label">INTERACTIONS</div><div class="metric-value">{fmt_num(summary['totals'].get('engagement', 0))}</div><div class="metric-note">赞藏评转合计</div></div>
      <div class="metric"><div class="metric-label">UTILITY RATIO</div><div class="metric-value">{summary['averages']['utility_ratio']:.2f}</div><div class="metric-note">收藏 / 点赞</div></div>
      <div class="metric"><div class="metric-label">FOLLOW / 1000</div><div class="metric-value">{summary['averages']['follows_per_1000']:.2f}</div><div class="metric-note">千次触达涨粉</div></div>
    </section>
    <section class="grid">
      <div class="card"><h2>板块触达排行</h2>{pillar_rows}</div>
      <div class="card"><h2>板块决策</h2><div class="pill-board">{decisions}</div></div>
    </section>
    <section class="grid">
      <div class="card"><h2>隐藏机会</h2><div class="next-list">{next_posts or '<div class="subtitle">需要更多涨粉/收藏数据判断隐藏机会。</div>'}</div></div>
      <div class="card"><h2>数据缺口</h2><p>{html.escape('、'.join(summary['data_gaps']) or '当前基础字段可用。')}</p></div>
    </section>
  </main>
</body>
</html>
"""
    path.write_text(html_doc, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--platform", default="小红书")
    parser.add_argument("--account", default="未命名账号")
    parser.add_argument("--output-dir", type=Path, default=Path("analysis-output"))
    parser.add_argument("--pillars-json", type=Path)
    parser.add_argument("--trend-file", type=Path)
    args = parser.parse_args()

    rows = read_rows(args.input)
    if not rows:
        raise SystemExit("No rows found.")
    field_map = map_fields(list(rows[0].keys()))
    enriched = enrich_rows(rows, field_map, load_pillars(args.pillars_json), load_trends(args.trend_file))
    summary = summarize(enriched, args.platform, args.account, args.input, field_map)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    write_json(summary, args.output_dir / "analysis-summary.json")
    write_markdown(summary, args.output_dir / "report-draft.md")
    write_html(summary, args.output_dir / "dashboard.html")
    theme_src = Path(__file__).resolve().parents[1] / "assets" / "dashboard-theme.css"
    theme_dst = args.output_dir / "dashboard-theme.css"
    if theme_src.exists():
        theme_dst.write_text(theme_src.read_text(encoding="utf-8"), encoding="utf-8")
    print(f"Wrote {args.output_dir}")


if __name__ == "__main__":
    main()
